from telethon import TelegramClient, events, sync
from telethon.tl.functions.users import GetFullUserRequest
import openpyxl
from telethon import functions, types
from RealUserExcelData import RealUserExcelData
import config
from SQLither import *
import asyncio
import random
import time

def get_client():
	api_id = 
	api_hash = 

	client = TelegramClient('session_name_1', api_id, api_hash)
	client.start()
	return client

def get_users():
	wb_obj = openpyxl.load_workbook(config.input_excel_users) 
	sheet_obj = wb_obj.active
	users = []
	print(sheet_obj.max_row)
	for i in range(sheet_obj.max_row):
		cell_obj_phone = sheet_obj.cell(row = i + 1, column = 1) 
		cell_obj_initials = sheet_obj.cell(row = i + 1, column = 2)
		# print(cell_obj_initials.value)
		# print(cell_obj_phone.value)
		if(cell_obj_phone.value == None):
			break
		phone = '+{}'.format(int(cell_obj_phone.value))
		initials = str(cell_obj_initials.value)
		# print("{} {}".format("+{}".format(int(cell_obj_phone.value)), str(cell_obj_initials.value)))
		users.append(RealUserExcelData(initials, phone))
	# print(users)
	return users

# # users = get_users()
client = get_client()

# client.send_message('+380631558792', 'Hello, friend!')

def is_real(imported_contact):
	return len(imported_contact.imported) != 0

def get_id(imported_contact):
	return imported_contact.users[0].id

def save_to_db(users, client):
	db_worker = SQLither(config.database_name)
	contacts = []
	for user in users:
		contacts.append(types.InputPhoneContact(
					client_id= random.randint(-10000000,10000000),
					phone=user.get_phone(),
					first_name=user.get_first_name(),
					last_name=user.get_last_name()
				))
	result = client(functions.contacts.ImportContactsRequest(
			contacts=contacts
	))	
	user_dic = dict()
	for user in users:
			user_dic[user.get_phone()[1:]] = user.get_initials()
	print("======================================")
	print("In telegram")
	print("======================================")
	phones = set()
	for user in result.users:
		print("{} - {}".format(user.phone, user_dic[user.phone]))
		phones.add(user.phone)
		db_worker.insert_real_user(user.id, user_dic[user.phone], "+{}".format(user.phone))

	
	print("======================================")
	print("Not in telegram")
	print("======================================")
	for user in users:
		if not user.get_phone()[1:] in phones:
			print("{} - {} {}".format(user.get_phone()[1:], user.get_first_name(), user.get_last_name()))
			db_worker.insert_non_exist_user(user.get_initials(), user.get_phone())



	# if i == 40:
	# 	time.sleep(300)
	# 	i = 0
	# if is_real(result):
	# 	db_worker.insert_real_user(get_id(result), user.get_initials(), user.get_phone())
	# else:
	# 	db_worker.insert_non_exist_user(user.get_initials(), user.get_phone())
	# i += 1
	
	db_worker.close()
save_to_db(get_users(), get_client())

# result = client(functions.contacts.ImportContactsRequest(
# 			contacts=[types.InputPhoneContact(
# 			client_id= random.randint(-10000000,10000000),
# 			phone='+380997011654',
# 			first_name='Іван',
# 			last_name='Кубарич'
# 			)]
# 		))

# print(result)